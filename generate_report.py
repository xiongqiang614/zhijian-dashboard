# -*- coding: utf-8 -*-
import openpyxl, json, os
from datetime import datetime

EXCEL_PATH = 'C:/Users/86135/Desktop/质量抽检表 (4).xlsx'
OUT_DIR = 'C:/Users/86135/WorkBuddy/2026-06-23-14-49-40'
JSON_PATH = os.path.join(OUT_DIR, 'chart_data.json')

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

def get_group(row, leader_col):
    if leader_col < 0: return "\u5ba2\u8bc9\u7ec4"
    l = str(row[leader_col] or "")
    if "\u7530\u654f" in l: return "\u7530\u654f\u7ec4"
    if "\u674e\u5e08" in l: return "\u674e\u5e08\u7ec4"
    if "\u7f57\u96c4\u5f3a" in l: return "\u7f57\u96c4\u5f3a\u7ec4"
    return "\u672a\u5206\u7ec4"

all_data = []
kf_count = 0
ks_count = 0

# Sheet configs: (sheet_idx, col_emp, col_date, col_checker, col_ct, col_et, col_lv, col_leader, col_score, col_month, src)
configs = [
    (4, 3, 0, 2, 6, 7, 8, 10, 11, 12, "kf"),  # \u5ba2\u670d
    (7, 3, 0, 2, 6, 7, 8, 10, 11, 12, "kf"),  # \u5ba2\u670d\u5f52\u6863
    (5, 3, 0, 2, 6, 7, 8, -1, 10, 11, "ks"), # \u5ba2\u8bc9
    (8, 3, 0, 2, 6, 7, 8, -1, 10, 11, "ks"), # \u5ba2\u8bc9\u5f52\u6863
]

for (si, ce, cd, cch, cct, cet, cl, clr, csc, cm, src) in configs:
    ws = wb.worksheets[si]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        try: emp = str(row[ce]).strip() if row[ce] else ""
        except: continue
        if not emp: continue
        try: sv = float(row[csc]) if row[csc] is not None else 0
        except: sv = 0
        month = str(row[cm]).strip() if row[cm] else ""
        all_data.append({
            "src": src, "emp": emp, "group": get_group(row, clr),
            "score": sv, "month": month,
            "checker": str(row[cch] or "")[:10],
            "ct": str(row[cct] or "")[:30],
            "et": str(row[cet] or "")[:30],
            "lv": str(row[cl] or "")[:20],
            "date": str(row[cd])[:10] if row[cd] else ""
        })
        if src == "kf": kf_count += 1
        else: ks_count += 1

# \u4e0d\u6ee1\u610f\u4e13\u9879 (sheet 10)
ws_bmy = wb.worksheets[10]
ci = {str(c.value): i for i, c in enumerate(ws_bmy[1])}
bmy = []
for row in ws_bmy.iter_rows(min_row=2, max_row=ws_bmy.max_row, values_only=True):
    emp = ""
    if "\u5458\u5de5" in ci and row[ci["\u5458\u5de5"]]:
        emp = str(row[ci["\u5458\u5de5"]]).strip()
    if not emp: continue
    bmy.append({
        "date": str(row[ci.get("\u8d28\u68c0\u65f6\u95f4", -1)] or "")[:10] if "\u8d28\u68c0\u65f6\u95f4" in ci else "",
        "emp": emp,
        "bigcat": str(row[ci.get("\u4e0d\u6ee1\u610f\u5927\u7c7b", -1)] or "").strip() if "\u4e0d\u6ee1\u610f\u5927\u7c7b" in ci else "",
    })

# ===== AGGREGATION =====
months = sorted(set(d["month"] for d in all_data if d["month"]))
cur_month = months[-1] if months else ""
emps = sorted(set(d["emp"] for d in all_data))
groups = sorted(set(d["group"] for d in all_data))
ets = sorted(set(d["et"] for d in all_data if d["et"]))
levels = sorted(set(d["lv"] for d in all_data if d["lv"]))
checkers = sorted(set(d["checker"] for d in all_data if d["checker"]))
cts = sorted(set(d["ct"] for d in all_data if d["ct"]))

def filtered(src_filter, month_filter=None, emp_filter=None, group_filter=None):
    data = all_data
    if src_filter != "all": data = [d for d in data if d["src"] == src_filter]
    if month_filter and month_filter != "all": data = [d for d in data if d["month"] == month_filter]
    if emp_filter and emp_filter != "all": data = [d for d in data if d["emp"] == emp_filter]
    if group_filter and group_filter != "all": data = [d for d in data if d["group"] == group_filter]
    return data

# Build full output
out = {
    "meta": {
        "gen_time": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "total": len(all_data), "kf": kf_count, "ks": ks_count, "bmy": len(bmy)
    },
    "dims": {
        "months": months, "cur_month": cur_month,
        "emps": emps, "groups": groups, "ets": ets,
        "levels": levels, "checkers": checkers, "cts": cts,
        "sources": ["kf", "ks"]
    },
    "kpi": {
        "total_deduct": len([d for d in all_data if d["score"] < 0]),
        "total_bonus": len([d for d in all_data if d["score"] > 0]),
        "total_errors": len([d for d in all_data if d["et"]]),
        "cur_month_deduct": len([d for d in all_data if d["month"]==cur_month and d["score"]<0]),
        "cur_month_bonus": len([d for d in all_data if d["month"]==cur_month and d["score"]>0]),
        "cur_month_errors": len([d for d in all_data if d["month"]==cur_month and d["et"]]),
        "cur_month_total": len([d for d in all_data if d["month"]==cur_month]),
        "monthly": {m: {
            "rec": len([d for d in all_data if d["month"]==m]),
            "deduct": len([d for d in all_data if d["month"]==m and d["score"]<0]),
            "bonus": len([d for d in all_data if d["month"]==m and d["score"]>0]),
            "errors": len([d for d in all_data if d["month"]==m and d["et"]])
        } for m in months},
    },
    "emp_stats": {e: {
        "group": next((d["group"] for d in all_data if d["emp"]==e), ""),
        "rec": len([d for d in all_data if d["emp"]==e]),
        "deduct": len([d for d in all_data if d["emp"]==e and d["score"]<0]),
        "bonus": len([d for d in all_data if d["emp"]==e and d["score"]>0]),
        "errors": len([d for d in all_data if d["emp"]==e and d["et"]])
    } for e in emps},
    "emp_et_dist": {e: {et: len([d for d in all_data if d["emp"]==e and d["et"]==et]) for et in ets} for e in emps},
    "emp_monthly": {e: {m: {
        "deduct": len([d for d in all_data if d["emp"]==e and d["month"]==m and d["score"]<0]),
        "bonus": len([d for d in all_data if d["emp"]==e and d["month"]==m and d["score"]>0])
    } for m in months} for e in emps},
    "group_stats": {g: {
        "rec": len([d for d in all_data if d["group"]==g]),
        "deduct": len([d for d in all_data if d["group"]==g and d["score"]<0]),
        "bonus": len([d for d in all_data if d["group"]==g and d["score"]>0]),
        "emp_cnt": len(set(d["emp"] for d in all_data if d["group"]==g))
    } for g in groups},
    "group_monthly": {g: {m: {
        "deduct": len([d for d in all_data if d["group"]==g and d["month"]==m and d["score"]<0]),
        "bonus": len([d for d in all_data if d["group"]==g and d["month"]==m and d["score"]>0])
    } for m in months} for g in groups},
    "et_stats": {et: len([d for d in all_data if d["et"]==et]) for et in ets},
    "et_monthly": {et: {m: len([d for d in all_data if d["et"]==et and d["month"]==m]) for m in months} for et in ets},
    "lv_data": {lv: len([d for d in all_data if d["lv"]==lv]) for lv in levels},
    "ct_data": {ct: len([d for d in all_data if d["ct"]==ct]) for ct in cts},
    "bmy": bmy,
    "bmy_stats": {c: sum(1 for d in bmy if d["bigcat"]==c) for c in sorted(set(d["bigcat"] for d in bmy if d["bigcat"]))},
    "raw": all_data,
}

with open(JSON_PATH, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

print(f"Done: {len(all_data)} records, {len(emps)} employees, {len(groups)} groups")
print(f"Months: {months[0]} ~ {months[-1]} ({len(months)} months)")
print(f"KF: {kf_count}, KS: {ks_count}, BMY: {len(bmy)}")
